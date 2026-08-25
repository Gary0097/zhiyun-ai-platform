# -*- coding: utf-8 -*-
"""Safe CSV/XLSX/XLS parser and builder for generic Data Core imports/exports."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

MAX_ROWS = 10000


def _clean_rows(rows: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        raise ValueError("文件中没有数据")
    headers = [str(value or "").strip() for value in rows[0]]
    active = [header for header in headers if header]
    if not active:
        raise ValueError("第一行必须是字段名称")
    if len(active) != len(set(active)):
        raise ValueError("字段名称不能重复")
    records = []
    for values in rows[1 : MAX_ROWS + 1]:
        if not any(value not in (None, "") for value in values):
            continue
        records.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header})
    return active, records


def _read_xls(content: bytes) -> list[list[Any]]:
    try:
        from xlrd import open_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("解析 .xls 需要安装 xlrd") from exc
    workbook = open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    return [sheet.row_values(index) for index in range(sheet.nrows)]


def _safe_sheet_title(value: str) -> str:
    """Sanitize a workbook sheet title to Excel's allowed characters."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "", value).strip()
    if not cleaned:
        cleaned = "Sheet1"
    return cleaned[:31]


def parse_table(filename: str, content: bytes) -> dict[str, Any]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        matrix = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        headers, rows = _clean_rows(matrix)
        return {"filename": filename, "sheet": None, "headers": headers, "rows": rows, "row_count": len(rows)}
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        headers, rows = _clean_rows([list(row) for row in sheet.iter_rows(values_only=True)])
        return {"filename": filename, "sheet": sheet.title, "headers": headers, "rows": rows, "row_count": len(rows)}
    if suffix == ".xls":
        headers, rows = _clean_rows(_read_xls(content))
        return {"filename": filename, "sheet": None, "headers": headers, "rows": rows, "row_count": len(rows)}
    raise ValueError("仅支持 .xlsx、.xls 和 .csv 文件")


def build_export_bytes(filename: str, headers: list[str], rows: list[dict[str, Any]]) -> tuple[bytes, str, str]:
    """Build export bytes and return (bytes, media_type, suggested_filename)."""
    base = Path(filename).stem
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(header) for header in headers])
        data = ("\ufeff" + buffer.getvalue()).encode("utf-8")
        return data, "text/csv; charset=utf-8", f"{base}.csv"
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(base)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{base}.xlsx"
