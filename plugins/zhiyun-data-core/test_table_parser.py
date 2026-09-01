# -*- coding: utf-8 -*-

import io
import sys
import types
import unittest
from unittest import mock

from table_parser import build_export_bytes, parse_table

try:
    import openpyxl  # noqa: F401
    _HAS_OPENPYXL = True
except ModuleNotFoundError:
    _HAS_OPENPYXL = False

_requires_openpyxl = unittest.skipUnless(
    _HAS_OPENPYXL, "openpyxl 未安装时跳过 xlsx 用例（CSV 路径不受影响）")


class TableParserTests(unittest.TestCase):
    def test_csv_is_parsed_for_any_dataset(self) -> None:
        result = parse_table("production.csv", "日期,部门,产量\n2026-08-22,一车间,120\n".encode())
        self.assertEqual(result["headers"], ["日期", "部门", "产量"])
        self.assertEqual(result["rows"][0]["部门"], "一车间")

    def test_duplicate_headers_are_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "不能重复"):
            parse_table("bad.csv", "日期,日期\n1,2\n".encode())

    def test_unknown_extension_is_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "仅支持.*\.xls"):
            parse_table("data.txt", b"a,b")

    @_requires_openpyxl
    def test_xlsx_parse_reads_first_sheet(self) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "订单"
        sheet.append(["order_no", "customer_name"])
        sheet.append(["A-1", "海川制造"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        result = parse_table("orders.xlsx", buffer.getvalue())
        self.assertEqual(result["sheet"], "订单")
        self.assertEqual(result["headers"], ["order_no", "customer_name"])
        self.assertEqual(result["rows"][0]["customer_name"], "海川制造")

    def test_xls_parse_requires_xlrd(self) -> None:
        # When the optional xlrd dependency is missing, .xls parsing must raise a
        # clear ValueError instead of leaking an ImportError as an HTTP 500.
        with mock.patch.dict(sys.modules, {"xlrd": None}):
            with self.assertRaisesRegex(ValueError, "需要安装 xlrd"):
                parse_table("legacy.xls", b"\xd0\xcf\x11\xe0")

    def test_xls_parse_handles_corrupt_workbook(self) -> None:
        # A malformed .xls byte stream must surface as a ValueError, not a 500.
        class FakeXLRDError(Exception):
            pass

        fake_xlrd = types.ModuleType("xlrd")
        fake_xlrd.biffh = types.SimpleNamespace(XLRDError=FakeXLRDError)

        def _boom(file_contents: bytes) -> None:
            raise FakeXLRDError("not a valid xls")

        fake_xlrd.open_workbook = _boom
        with mock.patch.dict(sys.modules, {"xlrd": fake_xlrd}):
            with self.assertRaisesRegex(ValueError, "无法解析 .xls"):
                parse_table("legacy.xls", b"\xd0\xcf\x11\xe0")

    def test_export_csv_bytes_are_bom_prefixed(self) -> None:
        data, media_type, suggested = build_export_bytes(
            "orders.csv",
            ["order_no", "customer_name"],
            [{"order_no": "A-1", "customer_name": "海川制造"}],
        )
        self.assertEqual(media_type, "text/csv; charset=utf-8")
        self.assertTrue(data.startswith(b"\xef\xbb\xbf"))
        self.assertEqual(suggested, "orders.csv")
        self.assertIn("海川制造".encode("utf-8"), data)

    @_requires_openpyxl
    def test_export_xlsx_bytes_round_trip(self) -> None:
        from openpyxl import load_workbook

        data, media_type, suggested = build_export_bytes(
            "orders.xlsx",
            ["order_no", "customer_name"],
            [{"order_no": "A-1", "customer_name": "海川制造"}],
        )
        self.assertEqual(media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(suggested, "orders.xlsx")
        self.assertTrue(data.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("order_no", "customer_name"))
        self.assertEqual(rows[1][1], "海川制造")


if __name__ == "__main__":
    unittest.main()
